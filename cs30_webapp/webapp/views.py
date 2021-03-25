from django.shortcuts import render
from django.http import HttpResponse
from webapp.forms import UserForm, EditForm, UploadForm
from django.contrib.auth import authenticate, login
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.template import RequestContext
from django.shortcuts import redirect
from django.urls import reverse
from django.contrib import messages
from openpyxl import load_workbook
import requests
import datetime


def home(request):
    return render(request, 'webapp/home.html')


def search(request):
    search = request.POST.get('search')

    '''
    As django_rest_framework cannot search IntegerFields as it can CharFields, this checks if the searched term
    is an integer and preforms a specific lookup for an entry with that ref_num to be appended to the list
    of results.
    '''
    specific_entry = []
    try:
        int(search)
        specific_entry = requests.get('http://cs30.herokuapp.com/api/carbon/' + search).json()
    except ValueError:
        pass

    entries = requests.get('http://cs30.herokuapp.com/api/carbon/search/' + search).json()

    # Performs the merging of possible entries, depending on which actually contain values.
    if specific_entry and entries:
        all_entries = entries.append(specific_entry)
    elif specific_entry and not entries:
        all_entries = [specific_entry]
    else:
        all_entries = entries

    # Checks the given date against the three possible formats they can take and then formats them.
    for entry in all_entries:
        for format in('%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d %H:%M:%S'):
            try:
                entry['other_info']['last_update'] = datetime.datetime.strptime(str(entry['other_info']['last_update']), format)
            except ValueError:
                pass

    return render(request, 'webapp/dbview.html', context = {'entries': all_entries, 'from': 'search'})





@login_required
def edit(request, refnum):
    entry = requests.get('http://cs30.herokuapp.com/api/carbon/' + refnum).json()

    # Checks the given date against the three possible formats they can take and then formats them.
    for format in('%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d %H:%M:%S'):
        try:
            entry['other_info']['last_update'] = datetime.datetime.strptime(str(entry['other_info']['last_update']), format)
        except ValueError:
            pass

    if request.method == 'POST':
        edit_form = EditForm(request.POST)

        if edit_form.is_valid():

            # Formats ther user given data into the correct format to be sent to the api.
            edit = {
                        'ref_num':refnum,
                        # While levels cannot be edited currently, I'm leaving this in as an example.
                        'navigation_info':{
                            'scope':request.POST.get('scope'),
                            'level1':request.POST.get('level1'),
                            'level2':request.POST.get('level2'),
                            'level3':request.POST.get('level3'),
                            'level4':request.POST.get('level4'),
                            'level5':request.POST.get('level5')
                            },
                        'calculation_info':{
                            'ef':float(request.POST.get('ef')),
                            'cu':request.POST.get('cu')
                            },
                        'other_info':{
                            'last_update':(datetime.datetime.now(tz=None)).__str__(),
                            'preference':int(request.POST.get('preference')),
                            'source':request.POST.get('source')
                            }
                    }


            # Attempts to send the data the api with a put request.
            try_edit = requests.put('http://cs30.herokuapp.com/api/carbon/' + refnum, json=edit)

            # Displays an error or success message depending on the status returned from the api.
            if try_edit.status_code == 200:
                messages.success(request, 'Edit successful!')
                return redirect(reverse('webapp:edit', kwargs={'refnum':refnum}))
            else:
                messages.error(request, 'Edit unsuccessful, please check edits are valid and try again.')
                return redirect(reverse('webapp:edit', kwargs={'refnum':refnum}))

        else:
            messages.error(request, 'Edit unsuccessful, please check edits are valid and try again.')
            return redirect(reverse('webapp:edit', kwargs={'refnum':refnum}))

    else:
        edit_form = EditForm()

    return render(request, 'webapp/edit.html', context={'edit_form': edit_form,'entry':entry})





@login_required
def add(request):
    if request.method == 'POST':

        # If the user has attempted to upload a file.
        if 'upload' in request.POST:
            upload_form = UploadForm()

            # Loads the .xlsx file supplied by the user
            wb = load_workbook(request.FILES['uploadfile'])
            ws = wb.active

            # Navigates through the .xlsx taking the values from each row and formatting them as required.
            for idx, row in enumerate(ws.iter_rows()):
                if idx == 0 or idx == 1:
                    continue
                else:
                    if ws.cell(row = idx, column = 1).value == None:
                        break
                    else:
                        edit = {
                                    'ref_num':ws.cell(row = idx+1, column = 1).value,
                                    'navigation_info':{
                                        'scope':ws.cell(row = idx+1, column = 2).value,
                                        'level1':ws.cell(row = idx+1, column = 3).value,
                                        'level2':ws.cell(row = idx+1, column = 4).value,
                                        'level3':ws.cell(row = idx+1, column = 5).value,
                                        'level4':ws.cell(row = idx+1, column = 6).value,
                                        'level5':ws.cell(row = idx+1, column = 7).value
                                        },
                                    'calculation_info':{
                                        'ef':ws.cell(row = idx+1, column = 10).value,
                                        'cu':ws.cell(row = idx+1, column = 9).value
                                        },
                                    'other_info':{
                                        'last_update':(ws.cell(row = idx+1, column = 12).value).__str__(),
                                        'preference':ws.cell(row = idx+1, column = 11).value,
                                        'source':ws.cell(row = idx+1, column = 8).value
                                        }
                                }

                        try_upload = requests.post('https://cs30.herokuapp.com/api/carbon', json=edit)

            messages.success(request, 'Upload successful!')
            return render(request, 'webapp/add.html', context={'upload_form': upload_form})

        # If the user has attempted to add a single entry
        elif 'save' in request.POST:
            upload_form = UploadForm(request.POST)

            if upload_form.is_valid():
                refnum = request.POST.get('ref_num')

                # Formats ther user given data into the correct format to be sent to the api.
                upload = {
                            'ref_num':int(refnum),
                            'navigation_info':{
                                'scope':request.POST.get('scope'),
                                'level1':request.POST.get('level1'),
                                'level2':request.POST.get('level2'),
                                'level3':request.POST.get('level3'),
                                'level4':request.POST.get('level4'),
                                'level5':request.POST.get('level5')
                                },
                            'calculation_info':{
                                'ef':float(request.POST.get('ef')),
                                'cu':request.POST.get('cu')
                                },
                            'other_info':{
                                'last_update':(datetime.datetime.now(tz=None)).__str__(),
                                'preference':int(request.POST.get('preference')),
                                'source':request.POST.get('source')
                                }
                        }

                # Sets any values left blank by the user as None for the api.
                for key, value in upload['navigation_info'].items():
                    if value == '':
                        upload['navigation_info'][key] = None

                try_upload = requests.post('https://cs30.herokuapp.com/api/carbon', json=upload)

                messages.success(request, 'Upload successful!')
                return render(request, 'webapp/add.html', context={'upload_form': upload_form})

            else:
                messages.error(request, 'Upload unsuccessful, please check inputs are valid and try again.')

    else:
        upload_form = UploadForm()

    return render(request, 'webapp/add.html', context={'upload_form': upload_form})





@login_required
def delete(request, refnum):
    if request.method == 'POST':
        requests.delete('http://cs30.herokuapp.com/api/carbon/' + refnum)
        return redirect(reverse('webapp:dbview'))
    return render(request, 'webapp/dbview.html', context = {'entries': entries})





@login_required
def dbview(request):
    # The url here will need to be made more general so it doesn't need to be changed based on host, I don't remember how to do that though
    entries = requests.get('http://cs30.herokuapp.com/api/carbon').json()

    # Checks the given date against the three possible formats they can take and then formats them.
    for entry in entries:
        for format in('%Y-%m-%dT%H:%M:%SZ', '%Y-%m-%dT%H:%M:%S.%fZ', '%Y-%m-%d %H:%M:%S'):
            try:
                entry['other_info']['last_update'] = datetime.datetime.strptime(str(entry['other_info']['last_update']), format)
            except ValueError:
                pass

    return render(request, 'webapp/dbview.html', context = {'entries': entries, 'from': 'dbview'})





def register(request):
    context = RequestContext(request)

    if request.method == 'POST':
        user_form = UserForm(request.POST)

        if user_form.is_valid():
            user = user_form.save()

            user.set_password(user.password)
            user.save()

            messages.success(request, 'Thank you for registering! Please wait for a staff member to activate your account.')
            return render(request, 'webapp/home.html')
        else:
            print(user_form.errors)

    else:
        user_form = UserForm()

    return render(request, 'webapp/register.html', context={'user_form': user_form})


def user_login(request):
    context = RequestContext(request)

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:
            # Ensures the account has been activated by staff before user is allowed to login.
            if user.is_staff:
                login(request, user)
                return redirect(reverse('webapp:home'))
            else:
                # An inactive account was used - no logging in!
                messages.error(request, 'Your account has not been activated, please contact a staff member.')
                return render(request, 'webapp/login.html')
        else:
            # Bad login details were provided. So we can't log the user in.

            messages.error(request, 'Username or password was incorrect, please try again.')
            return render(request, 'webapp/login.html')

    else:

        return render(request, 'webapp/login.html')


@login_required
def user_logout(request):
    # Since we know the user is logged in, we can now just log them out.
    logout(request)
    return HttpResponseRedirect("/webapp/")
